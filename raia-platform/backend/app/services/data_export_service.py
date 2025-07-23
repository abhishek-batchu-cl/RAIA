"""
Data Export Service
Comprehensive data export functionality supporting PDF, Excel, and CSV formats
"""

import asyncio
import csv
import json
import uuid
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import structlog

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.platypus import PageBreak
from reportlab.lib import colors

logger = structlog.get_logger(__name__)


class DataExportService:
    """
    Comprehensive data export service supporting multiple formats
    """
    
    def __init__(self, base_export_path: str = "/tmp/exports"):
        self.base_export_path = Path(base_export_path)
        self.base_export_path.mkdir(parents=True, exist_ok=True)
        self.logger = logger.bind(service="data_export")
        
        # Export format handlers
        self.format_handlers = {
            "pdf": self._export_to_pdf,
            "excel": self._export_to_excel,
            "csv": self._export_to_csv
        }
    
    async def export_evaluation_results(
        self,
        evaluation_data: Dict[str, Any],
        export_format: str,
        export_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Export evaluation results in specified format
        
        Args:
            evaluation_data: Evaluation data to export
            export_format: Export format (pdf, excel, csv)
            export_options: Additional export options
            
        Returns:
            Export job details with file path and metadata
        """
        export_id = str(uuid.uuid4())
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        
        if export_format not in self.format_handlers:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        self.logger.info(
            "Starting evaluation export",
            export_id=export_id,
            format=export_format,
            data_size=len(str(evaluation_data))
        )
        
        try:
            # Generate filename
            filename = f"evaluation_report_{timestamp}.{export_format}"
            file_path = self.base_export_path / filename
            
            # Call appropriate format handler
            handler = self.format_handlers[export_format]
            await handler(evaluation_data, file_path, export_options or {})
            
            # Get file size
            file_size = file_path.stat().st_size
            
            return {
                "export_id": export_id,
                "filename": filename,
                "file_path": str(file_path),
                "file_size_bytes": file_size,
                "format": export_format,
                "created_at": datetime.utcnow().isoformat(),
                "expires_at": (datetime.utcnow() + timedelta(days=7)).isoformat()
            }
            
        except Exception as e:
            self.logger.error(
                "Export failed",
                export_id=export_id,
                error=str(e)
            )
            raise
    
    async def export_model_performance_report(
        self,
        model_data: Dict[str, Any],
        performance_metrics: Dict[str, Any],
        export_format: str,
        export_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export comprehensive model performance report"""
        
        export_data = {
            "report_type": "Model Performance Report",
            "model_info": model_data,
            "performance_metrics": performance_metrics,
            "generated_at": datetime.utcnow().isoformat(),
            "sections": [
                "Model Overview",
                "Performance Metrics",
                "Evaluation History",
                "Recommendations"
            ]
        }
        
        return await self.export_evaluation_results(
            evaluation_data=export_data,
            export_format=export_format,
            export_options=export_options
        )
    
    async def export_fairness_analysis_report(
        self,
        fairness_data: Dict[str, Any],
        export_format: str,
        export_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export fairness analysis report"""
        
        export_data = {
            "report_type": "Fairness Analysis Report",
            "fairness_metrics": fairness_data,
            "generated_at": datetime.utcnow().isoformat(),
            "sections": [
                "Fairness Overview",
                "Bias Detection Results",
                "Demographic Parity Analysis",
                "Mitigation Recommendations"
            ]
        }
        
        return await self.export_evaluation_results(
            evaluation_data=export_data,
            export_format=export_format,
            export_options=export_options
        )
    
    async def export_data_drift_report(
        self,
        drift_data: Dict[str, Any],
        export_format: str,
        export_options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Export data drift analysis report"""
        
        export_data = {
            "report_type": "Data Drift Analysis Report",
            "drift_analysis": drift_data,
            "generated_at": datetime.utcnow().isoformat(),
            "sections": [
                "Drift Detection Summary",
                "Feature Drift Analysis",
                "Distribution Comparisons",
                "Impact Assessment"
            ]
        }
        
        return await self.export_evaluation_results(
            evaluation_data=export_data,
            export_format=export_format,
            export_options=export_options
        )
    
    async def _export_to_pdf(
        self,
        data: Dict[str, Any],
        file_path: Path,
        options: Dict[str, Any]
    ):
        """Export data to PDF format"""
        
        # Create PDF document
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=HexColor('#2E86C1')
        )
        story.append(Paragraph(data.get("report_type", "RAIA Platform Report"), title_style))
        story.append(Spacer(1, 20))
        
        # Metadata
        meta_style = styles['Normal']
        story.append(Paragraph(f"Generated: {data.get('generated_at', 'N/A')}", meta_style))
        story.append(Paragraph(f"Report ID: {str(uuid.uuid4())}", meta_style))
        story.append(Spacer(1, 30))
        
        # Executive Summary
        if 'summary' in data:
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            story.append(Paragraph(data['summary'], styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Model Information
        if 'model_info' in data:
            story.append(Paragraph("Model Information", styles['Heading2']))
            model_info = data['model_info']
            
            model_table_data = [
                ['Model Name', model_info.get('name', 'N/A')],
                ['Model Type', model_info.get('type', 'N/A')],
                ['Version', model_info.get('version', 'N/A')],
                ['Created', model_info.get('created_at', 'N/A')]
            ]
            
            model_table = Table(model_table_data, colWidths=[2*inch, 4*inch])
            model_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(model_table)
            story.append(Spacer(1, 20))
        
        # Performance Metrics
        if 'performance_metrics' in data:
            story.append(Paragraph("Performance Metrics", styles['Heading2']))
            metrics = data['performance_metrics']
            
            metrics_data = [['Metric', 'Value', 'Benchmark']]
            for metric, value in metrics.items():
                if isinstance(value, (int, float)):
                    benchmark = "Good" if value > 0.8 else "Needs Improvement"
                    metrics_data.append([metric.replace('_', ' ').title(), f"{value:.4f}", benchmark])
            
            metrics_table = Table(metrics_data, colWidths=[2*inch, 1.5*inch, 2*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 20))
        
        # Fairness Metrics
        if 'fairness_metrics' in data:
            story.append(Paragraph("Fairness Analysis", styles['Heading2']))
            fairness = data['fairness_metrics']
            
            story.append(Paragraph(f"Overall Fairness Score: {fairness.get('overall_score', 'N/A')}", styles['Normal']))
            
            if 'bias_metrics' in fairness:
                bias_data = [['Protected Attribute', 'Bias Score', 'Status']]
                for attr, score in fairness['bias_metrics'].items():
                    status = "Fair" if score < 0.1 else "Biased"
                    bias_data.append([attr.replace('_', ' ').title(), f"{score:.4f}", status])
                
                bias_table = Table(bias_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
                bias_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(bias_table)
            story.append(Spacer(1, 20))
        
        # Sections
        if 'sections' in data:
            story.append(PageBreak())
            story.append(Paragraph("Report Sections", styles['Heading2']))
            for i, section in enumerate(data['sections'], 1):
                story.append(Paragraph(f"{i}. {section}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Footer
        story.append(Spacer(1, 50))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph("Generated by RAIA Platform - Enterprise AI Evaluation", footer_style))
        
        # Build PDF
        doc.build(story)
    
    async def _export_to_excel(
        self,
        data: Dict[str, Any],
        file_path: Path,
        options: Dict[str, Any]
    ):
        """Export data to Excel format"""
        
        workbook = Workbook()
        
        # Remove default worksheet
        workbook.remove(workbook.active)
        
        # Create Summary worksheet
        summary_ws = workbook.create_sheet("Summary")
        self._create_excel_summary_sheet(summary_ws, data)
        
        # Create Model Info worksheet
        if 'model_info' in data:
            model_ws = workbook.create_sheet("Model Information")
            self._create_excel_model_sheet(model_ws, data['model_info'])
        
        # Create Performance Metrics worksheet
        if 'performance_metrics' in data:
            perf_ws = workbook.create_sheet("Performance Metrics")
            self._create_excel_performance_sheet(perf_ws, data['performance_metrics'])
        
        # Create Fairness Analysis worksheet
        if 'fairness_metrics' in data:
            fair_ws = workbook.create_sheet("Fairness Analysis")
            self._create_excel_fairness_sheet(fair_ws, data['fairness_metrics'])
        
        # Create Data worksheet for raw data
        if 'raw_data' in data:
            data_ws = workbook.create_sheet("Raw Data")
            self._create_excel_data_sheet(data_ws, data['raw_data'])
        
        # Save workbook
        workbook.save(str(file_path))
    
    def _create_excel_summary_sheet(self, worksheet, data):
        """Create Excel summary sheet"""
        
        # Title
        worksheet['A1'] = data.get('report_type', 'RAIA Platform Report')
        worksheet['A1'].font = Font(size=16, bold=True, color="2E86C1")
        worksheet.merge_cells('A1:D1')
        
        # Metadata
        worksheet['A3'] = 'Generated:'
        worksheet['B3'] = data.get('generated_at', 'N/A')
        worksheet['A4'] = 'Report ID:'
        worksheet['B4'] = str(uuid.uuid4())
        
        # Summary metrics
        row = 6
        worksheet[f'A{row}'] = 'Key Metrics'
        worksheet[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if 'performance_metrics' in data:
            for metric, value in data['performance_metrics'].items():
                worksheet[f'A{row}'] = metric.replace('_', ' ').title()
                worksheet[f'B{row}'] = value
                row += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_excel_model_sheet(self, worksheet, model_info):
        """Create Excel model information sheet"""
        
        worksheet['A1'] = 'Model Information'
        worksheet['A1'].font = Font(size=14, bold=True)
        
        row = 3
        for key, value in model_info.items():
            worksheet[f'A{row}'] = key.replace('_', ' ').title()
            worksheet[f'B{row}'] = value
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_excel_performance_sheet(self, worksheet, performance_metrics):
        """Create Excel performance metrics sheet"""
        
        worksheet['A1'] = 'Performance Metrics'
        worksheet['A1'].font = Font(size=14, bold=True)
        
        # Headers
        worksheet['A3'] = 'Metric'
        worksheet['B3'] = 'Value'
        worksheet['C3'] = 'Benchmark'
        
        # Style headers
        for cell in ['A3', 'B3', 'C3']:
            worksheet[cell].font = Font(bold=True)
            worksheet[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        row = 4
        for metric, value in performance_metrics.items():
            worksheet[f'A{row}'] = metric.replace('_', ' ').title()
            worksheet[f'B{row}'] = value
            if isinstance(value, (int, float)):
                benchmark = "Good" if value > 0.8 else "Needs Improvement"
                worksheet[f'C{row}'] = benchmark
            row += 1
    
    def _create_excel_fairness_sheet(self, worksheet, fairness_metrics):
        """Create Excel fairness analysis sheet"""
        
        worksheet['A1'] = 'Fairness Analysis'
        worksheet['A1'].font = Font(size=14, bold=True)
        
        worksheet['A3'] = 'Overall Fairness Score:'
        worksheet['B3'] = fairness_metrics.get('overall_score', 'N/A')
        
        if 'bias_metrics' in fairness_metrics:
            worksheet['A5'] = 'Bias Metrics'
            worksheet['A5'].font = Font(bold=True)
            
            # Headers
            worksheet['A7'] = 'Protected Attribute'
            worksheet['B7'] = 'Bias Score'
            worksheet['C7'] = 'Status'
            
            row = 8
            for attr, score in fairness_metrics['bias_metrics'].items():
                worksheet[f'A{row}'] = attr.replace('_', ' ').title()
                worksheet[f'B{row}'] = score
                status = "Fair" if score < 0.1 else "Biased"
                worksheet[f'C{row}'] = status
                row += 1
    
    def _create_excel_data_sheet(self, worksheet, raw_data):
        """Create Excel raw data sheet"""
        
        worksheet['A1'] = 'Raw Data'
        worksheet['A1'].font = Font(size=14, bold=True)
        
        if isinstance(raw_data, list) and raw_data:
            # Convert to DataFrame for easier handling
            df = pd.DataFrame(raw_data)
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                worksheet.cell(row=3, column=col_idx, value=column)
                worksheet.cell(row=3, column=col_idx).font = Font(bold=True)
            
            # Write data
            for row_idx, row_data in enumerate(df.values, 4):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    async def _export_to_csv(
        self,
        data: Dict[str, Any],
        file_path: Path,
        options: Dict[str, Any]
    ):
        """Export data to CSV format"""
        
        # For CSV, we'll export the most relevant tabular data
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write metadata header
            writer.writerow(['RAIA Platform Export'])
            writer.writerow(['Report Type', data.get('report_type', 'N/A')])
            writer.writerow(['Generated', data.get('generated_at', 'N/A')])
            writer.writerow([])  # Empty row
            
            # Performance metrics
            if 'performance_metrics' in data:
                writer.writerow(['Performance Metrics'])
                writer.writerow(['Metric', 'Value'])
                
                for metric, value in data['performance_metrics'].items():
                    writer.writerow([metric.replace('_', ' ').title(), value])
                
                writer.writerow([])  # Empty row
            
            # Fairness metrics
            if 'fairness_metrics' in data:
                writer.writerow(['Fairness Analysis'])
                writer.writerow(['Overall Score', data['fairness_metrics'].get('overall_score', 'N/A')])
                writer.writerow([])
                
                if 'bias_metrics' in data['fairness_metrics']:
                    writer.writerow(['Protected Attribute', 'Bias Score', 'Status'])
                    for attr, score in data['fairness_metrics']['bias_metrics'].items():
                        status = "Fair" if score < 0.1 else "Biased"
                        writer.writerow([attr.replace('_', ' ').title(), score, status])
                
                writer.writerow([])
            
            # Raw data
            if 'raw_data' in data and isinstance(data['raw_data'], list):
                writer.writerow(['Raw Data'])
                
                if data['raw_data']:
                    # Assume first item has the keys for headers
                    if isinstance(data['raw_data'][0], dict):
                        headers = list(data['raw_data'][0].keys())
                        writer.writerow(headers)
                        
                        for item in data['raw_data']:
                            writer.writerow([item.get(key, '') for key in headers])
    
    async def get_export_status(self, export_id: str) -> Dict[str, Any]:
        """Get export job status"""
        
        # In a real implementation, this would check database for export job status
        return {
            "export_id": export_id,
            "status": "completed",
            "progress_percentage": 100.0,
            "created_at": datetime.utcnow().isoformat(),
            "completed_at": datetime.utcnow().isoformat()
        }
    
    async def cleanup_expired_exports(self):
        """Clean up expired export files"""
        
        cutoff_date = datetime.utcnow() - timedelta(days=7)
        
        for file_path in self.base_export_path.iterdir():
            if file_path.is_file():
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                if file_time < cutoff_date:
                    file_path.unlink()
                    self.logger.info(
                        "Cleaned up expired export",
                        filename=file_path.name,
                        age_days=(datetime.utcnow() - file_time).days
                    )